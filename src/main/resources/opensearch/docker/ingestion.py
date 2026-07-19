#!/usr/bin/env python3

import csv
import hashlib
import os
import pathlib

from bs4 import BeautifulSoup
from docx import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter
from openai import OpenAI
from opensearchpy import OpenSearch
from openpyxl import load_workbook
from pypdf import PdfReader

BASE_URL = os.environ.get(
    "MODEL_RUNNER_BASE_URL",
    "http://model-runner.docker.internal/engines/llama.cpp/v1/",
)
EMBED_MODEL = os.environ.get("MODEL_RUNNER_LLM_EMBEDDING", "ai/nomic-embed-text-v2-moe")
DATA_DIR = os.environ.get("DATA_DIR", "/app/data")
OPENSEARCH_URL = os.environ.get("OPENSEARCH_URL", "http://opensearch:9200")
INDEX_NAME = os.environ.get("INDEX_NAME", "basaltrock-knowledge-base-default-index")

client = OpenAI(base_url=BASE_URL, api_key="dummy")
os_client = OpenSearch([OPENSEARCH_URL], use_ssl=False, verify_certs=False)
text_splitter = RecursiveCharacterTextSplitter(
    chunk_size=400,
    chunk_overlap=50,
)


def _embed(texts: list[str], batch_size: int = 10) -> list[list[float]]:
    embeddings = []
    for i in range(0, len(texts), batch_size):
        batch = texts[i:i + batch_size]
        resp = client.embeddings.create(model=EMBED_MODEL, input=batch)
        embeddings.extend([d.embedding for d in resp.data])
    return embeddings


def _extract_text(f: pathlib.Path) -> str:
    suffix = f.suffix.lower()

    if suffix == ".pdf":
        reader = PdfReader(f)
        return "\n".join(page.extract_text() for page in reader.pages)

    if suffix in [".doc", ".docx"]:
        doc = Document(f)
        return "\n".join(p.text for p in doc.paragraphs)

    if suffix in [".xls", ".xlsx"]:
        wb = load_workbook(f, data_only=True)
        lines = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                lines.append("\t".join(str(cell) if cell is not None else "" for cell in row))
        return "\n".join(lines)

    if suffix == ".csv":
        with open(f, encoding="utf-8", errors="ignore") as csvfile:
            reader = csv.reader(csvfile)
            return "\n".join("\t".join(row) for row in reader)

    if suffix in [".html", ".htm"]:
        raw_text = f.read_text(encoding="utf-8", errors="ignore")
        soup = BeautifulSoup(raw_text, "html.parser")
        return soup.get_text(separator="\n", strip=True)

    return f.read_text(encoding="utf-8", errors="ignore").strip()


def _load_chunks(folder: str) -> list[tuple[str, str]]:
    chunks = []
    patterns = ["*.txt", "*.html", "*.htm", "*.md", "*.pdf", "*.doc", "*.docx", "*.xls", "*.xlsx", "*.csv"]
    for pattern in patterns:
        for f in sorted(pathlib.Path(folder).rglob(pattern)):
            if f.name.startswith("._"):
                continue
            try:
                text = _extract_text(f)
            except Exception as e:
                print(f"  {f.name}: SKIPPED ({e})")
                continue
            chunk_texts = text_splitter.split_text(text)
            print(f"  {f.name}: {len(chunk_texts)} chunks")
            for chunk in chunk_texts:
                chunks.append((f.name, chunk))
    return chunks


raw = _load_chunks(DATA_DIR)

if not raw:
    print(f"No files found in '{DATA_DIR}'. Nothing to ingest.")
    exit(0)

if not os_client.indices.exists(INDEX_NAME):
    print(f"Creating index '{INDEX_NAME}'…")
    os_client.indices.create(
        index=INDEX_NAME,
        body={
            "settings": {"index.knn": True},
            "mappings": {
                "properties": {
                    "file": {"type": "keyword"},
                    "text": {"type": "text"},
                    "vec": {
                        "type": "knn_vector",
                        "dimension": 768,
                        "method": {"name": "hnsw", "engine": "lucene"}
                    }
                }
            }
        }
    )

print(f"Ingesting {len(raw)} chunks into '{INDEX_NAME}'…")

existing_docs = {}
result = os_client.search(index=INDEX_NAME, body={"query": {"match_all": {}}, "size": 10000})
for hit in result["hits"]["hits"]:
    key = (hit["_source"]["file"], hit["_source"]["text"])
    existing_docs[key] = hit["_id"]

fnames, texts = zip(*raw)
embeddings = _embed(list(texts))

for fname, text, vec in zip(fnames, texts, embeddings):
    key = (fname, text)
    doc_id = existing_docs.get(key, hashlib.md5((fname + text).encode()).hexdigest())
    os_client.index(
        index=INDEX_NAME,
        id=doc_id,
        body={"file": fname, "text": text, "vec": vec}
    )

os_client.indices.refresh(index=INDEX_NAME)
print(f"✓ {len(embeddings)} embeddings saved to OpenSearch index '{INDEX_NAME}'")